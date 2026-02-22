"""Export service for generating reports in multiple formats."""
from io import BytesIO
from typing import Optional

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.schemas.schemas import PropertyResponse, OpportunityScoreResponse


class ExportService:
    """Service for exporting property analysis to multiple formats."""

    def __init__(self):
        """Initialize export service."""
        pass

    def export_to_json(
        self,
        property_data: PropertyResponse,
        score_data: OpportunityScoreResponse,
        precedents: list[dict] = None,
        scenarios: list[dict] = None,
    ) -> dict:
        """
        Export analysis to JSON format.

        Args:
            property_data: Property information
            score_data: Opportunity score
            precedents: Planning precedents
            scenarios: Financial scenarios

        Returns:
            Dictionary ready for JSON serialization
        """
        return {
            "property": property_data.dict(),
            "score": score_data.dict(),
            "precedents": precedents or [],
            "scenarios": scenarios or [],
            "metadata": {
                "generated_at": __import__("datetime").datetime.now().isoformat(),
                "version": "1.0",
            },
        }

    def export_to_excel(
        self,
        property_data: PropertyResponse,
        score_data: OpportunityScoreResponse,
        precedents: list[dict] = None,
        scenarios: list[dict] = None,
    ) -> BytesIO:
        """
        Export analysis to Excel format.

        Args:
            property_data: Property information
            score_data: Opportunity score
            precedents: Planning precedents
            scenarios: Financial scenarios

        Returns:
            BytesIO object with Excel file
        """
        workbook = Workbook()
        workbook.remove(workbook.active)

        # Define styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=12)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Property Sheet
        ws_property = workbook.create_sheet("Property")
        ws_property["A1"] = "Property Analysis"
        ws_property["A1"].font = title_font

        # Property details
        row = 3
        ws_property[f"A{row}"] = "Address"
        ws_property[f"B{row}"] = property_data.address
        row += 1
        ws_property[f"A{row}"] = "Postcode"
        ws_property[f"B{row}"] = property_data.postcode
        row += 1
        ws_property[f"A{row}"] = "Current Use"
        ws_property[f"B{row}"] = property_data.historic_use or "N/A"
        row += 1
        ws_property[f"A{row}"] = "Listed Status"
        ws_property[f"B{row}"] = property_data.listed_status or "Not Listed"
        row += 1
        ws_property[f"A{row}"] = "Area (m²)"
        ws_property[f"B{row}"] = property_data.area_sqm or "N/A"

        # Score Sheet
        ws_score = workbook.create_sheet("Score")
        ws_score["A1"] = "Opportunity Score"
        ws_score["A1"].font = title_font

        row = 3
        ws_score[f"A{row}"] = "Total Score"
        ws_score[f"B{row}"] = score_data.score_total
        row += 1
        ws_score[f"A{row}"] = "Risk Score"
        ws_score[f"B{row}"] = score_data.score_risk
        row += 1
        ws_score[f"A{row}"] = "Confidence Score"
        ws_score[f"B{row}"] = score_data.score_confidence

        # Score breakdown
        if score_data.score_breakdown:
            row += 3
            ws_score[f"A{row}"] = "Breakdown"
            ws_score[f"A{row}"].font = title_font
            row += 1

            breakdown = score_data.score_breakdown
            ws_score[f"A{row}"] = "Precedents Found"
            ws_score[f"B{row}"] = breakdown.precedents_found
            row += 1
            ws_score[f"A{row}"] = "Approved"
            ws_score[f"B{row}"] = breakdown.approved
            row += 1
            ws_score[f"A{row}"] = "Refused"
            ws_score[f"B{row}"] = breakdown.refused
            row += 1
            ws_score[f"A{row}"] = "Approval Rate"
            ws_score[f"B{row}"] = f"{breakdown.approval_rate * 100:.1f}%"

        # Precedents Sheet
        if precedents:
            ws_precedents = workbook.create_sheet("Precedents")
            headers = ["Reference", "Proposal", "Decision", "Distance (m)", "Date"]
            for col, header in enumerate(headers, 1):
                cell = ws_precedents.cell(row=1, column=col)
                cell.value = header
                cell.fill = section_fill
                cell.font = header_font

            for row_idx, precedent in enumerate(precedents, 2):
                ws_precedents.cell(row=row_idx, column=1).value = precedent.get("reference", "")
                ws_precedents.cell(row=row_idx, column=2).value = precedent.get("proposal", "")[:50]
                ws_precedents.cell(row=row_idx, column=3).value = precedent.get("decision", "")
                ws_precedents.cell(row=row_idx, column=4).value = precedent.get("distance_m", 0)
                ws_precedents.cell(row=row_idx, column=5).value = precedent.get("date_decided", "")

        # Scenarios Sheet
        if scenarios:
            ws_scenarios = workbook.create_sheet("Scenarios")
            headers = ["Name", "Strategy", "Dev Cost", "ROI %", "Estimated Profit"]
            for col, header in enumerate(headers, 1):
                cell = ws_scenarios.cell(row=1, column=col)
                cell.value = header
                cell.fill = section_fill
                cell.font = header_font

            for row_idx, scenario in enumerate(scenarios, 2):
                ws_scenarios.cell(row=row_idx, column=1).value = scenario.get("name", "")
                ws_scenarios.cell(row=row_idx, column=2).value = scenario.get("strategy", "")
                ws_scenarios.cell(row=row_idx, column=3).value = scenario.get("development_cost", 0)
                ws_scenarios.cell(row=row_idx, column=4).value = f"{scenario.get('roi_percent', 0):.1f}%"
                ws_scenarios.cell(row=row_idx, column=5).value = f"£{scenario.get('estimated_profit', 0):,.0f}"

        # Adjust column widths
        for ws in workbook.sheetnames:
            worksheet = workbook[ws]
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def export_to_pdf(
        self,
        property_data: PropertyResponse,
        score_data: OpportunityScoreResponse,
        precedents: list[dict] = None,
        scenarios: list[dict] = None,
    ) -> BytesIO:
        """
        Export analysis to PDF format.

        Note: Requires weasyprint - for now returns HTML that can be rendered.

        Args:
            property_data: Property information
            score_data: Opportunity score
            precedents: Planning precedents
            scenarios: Financial scenarios

        Returns:
            BytesIO object with PDF file
        """
        try:
            from weasyprint import HTML, CSS
        except ImportError:
            logger.warning("weasyprint not installed - returning HTML instead")
            return self._generate_html_report(property_data, score_data, precedents, scenarios)

        # Generate HTML
        html_content = self._generate_html_report(property_data, score_data, precedents, scenarios)

        # Convert HTML to PDF
        try:
            pdf_bytes = HTML(string=html_content).write_pdf()
            return BytesIO(pdf_bytes)
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            return BytesIO(html_content.encode())

    def _generate_html_report(
        self,
        property_data: PropertyResponse,
        score_data: OpportunityScoreResponse,
        precedents: list[dict] = None,
        scenarios: list[dict] = None,
    ) -> str:
        """Generate HTML report content."""
        breakdown = score_data.score_breakdown or {}
        approval_rate = breakdown.approval_rate * 100 if breakdown.approval_rate else 0

        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>Use-Change Hunter Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #1F4E78; border-bottom: 2px solid #1F4E78; padding-bottom: 10px; }}
                h2 {{ color: #2E5090; margin-top: 20px; }}
                table {{ border-collapse: collapse; width: 100%; margin: 15px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #1F4E78; color: white; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .score-box {{ background-color: #D9E1F2; padding: 15px; border-radius: 5px; margin: 10px 0; }}
                .score-value {{ font-size: 24px; font-weight: bold; color: #1F4E78; }}
                .approval-rate {{ color: green; font-weight: bold; }}
            </style>
        </head>
        <body>
            <h1>Use-Change Hunter - Property Analysis Report</h1>

            <h2>Property Details</h2>
            <p><strong>Address:</strong> {property_data.address}</p>
            <p><strong>Postcode:</strong> {property_data.postcode}</p>
            <p><strong>Current Use:</strong> {property_data.historic_use or 'N/A'}</p>
            <p><strong>Listed Status:</strong> {property_data.listed_status or 'Not Listed'}</p>
            <p><strong>Area:</strong> {property_data.area_sqm or 'N/A'} m²</p>

            <h2>Opportunity Score</h2>
            <div class="score-box">
                <p><strong>Total Score:</strong> <span class="score-value">{score_data.score_total}</span> / 100</p>
                <p><strong>Risk Score:</strong> {score_data.score_risk} / 100</p>
                <p><strong>Confidence Score:</strong> {score_data.score_confidence} / 100</p>
            </div>

            <h2>Score Breakdown</h2>
            <table>
                <tr>
                    <th>Metric</th>
                    <th>Value</th>
                </tr>
                <tr>
                    <td>Precedents Found</td>
                    <td>{breakdown.get('precedents_found', 0)}</td>
                </tr>
                <tr>
                    <td>Approved Applications</td>
                    <td>{breakdown.get('approved', 0)}</td>
                </tr>
                <tr>
                    <td>Refused Applications</td>
                    <td>{breakdown.get('refused', 0)}</td>
                </tr>
                <tr>
                    <td>Approval Rate</td>
                    <td class="approval-rate">{approval_rate:.1f}%</td>
                </tr>
            </table>

            <h2>Generated Report</h2>
            <p>Report generated by Use-Change Hunter on {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </body>
        </html>
        """
        return html


# Singleton instance
_export_service: Optional[ExportService] = None


def get_export_service() -> ExportService:
    """Get or create export service singleton."""
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service
